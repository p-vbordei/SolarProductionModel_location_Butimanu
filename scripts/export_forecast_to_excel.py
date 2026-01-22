"""
Export intraday forecast CSV files to Excel format

IMPORTANT: All timestamps in Excel output are in CET/CEST (Europe/Berlin) timezone
"""
import pandas as pd
import os
from datetime import datetime
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sys

def export_forecast_to_excel(output_path=None):
    """
    Export the latest intraday forecast CSV files to Excel
    
    Args:
        output_path: Path for Excel file (optional, defaults to data_output/intraday)
    """
    try:
        # Set data directory
        data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data_output', 'intraday')
        
        # Find latest CSV files
        csv_files = [f for f in os.listdir(data_dir) if f.endswith('.csv') and 'cm_forecast_intraday' in f]
        
        # Separate 15min and 1hour files
        files_15min = [f for f in csv_files if '15min' in f]
        files_1hour = [f for f in csv_files if '1hour' in f]
        
        if not files_15min or not files_1hour:
            print("Error: Could not find forecast CSV files")
            print(f"Looking in directory: {data_dir}")
            print(f"Found {len(files_15min)} 15-minute files and {len(files_1hour)} hourly files")
            return None
        
        # Sort by modification time to get the truly latest files
        files_15min.sort(key=lambda x: os.path.getmtime(os.path.join(data_dir, x)))
        files_1hour.sort(key=lambda x: os.path.getmtime(os.path.join(data_dir, x)))
        
        # Use latest files
        latest_15min = files_15min[-1]
        latest_1hour = files_1hour[-1]
        
        print(f"Loading 15-minute data: {latest_15min}")
        print(f"Loading hourly data: {latest_1hour}")
        
        # Read CSV files (skip comment lines starting with #)
        df_15min = pd.read_csv(os.path.join(data_dir, latest_15min), comment='#')
        df_1hour = pd.read_csv(os.path.join(data_dir, latest_1hour), comment='#')

        # Convert MWh to kWh (multiply by 1000)
        energy_columns = ['energy_mwh', 'energy_q10_mwh', 'energy_q25_mwh', 'energy_q50_mwh', 'energy_q75_mwh', 'energy_q90_mwh']
        for col in energy_columns:
            if col in df_15min.columns:
                df_15min[col] = df_15min[col] * 1000
            if col in df_1hour.columns:
                df_1hour[col] = df_1hour[col] * 1000

        # Rename columns from MWh to kWh
        rename_map = {
            'energy_mwh': 'energy_kwh',
            'energy_q10_mwh': 'energy_q10_kwh',
            'energy_q25_mwh': 'energy_q25_kwh',
            'energy_q50_mwh': 'energy_q50_kwh',
            'energy_q75_mwh': 'energy_q75_kwh',
            'energy_q90_mwh': 'energy_q90_kwh'
        }
        df_15min = df_15min.rename(columns=rename_map)
        df_1hour = df_1hour.rename(columns=rename_map)

        # Fix column names if timestamp is unnamed
        if 'Unnamed: 0' in df_15min.columns:
            df_15min = df_15min.rename(columns={'Unnamed: 0': 'timestamp'})
        elif df_15min.columns[0] == '' or pd.isna(df_15min.columns[0]):
            df_15min = df_15min.rename(columns={df_15min.columns[0]: 'timestamp'})
            
        if 'Unnamed: 0' in df_1hour.columns:
            df_1hour = df_1hour.rename(columns={'Unnamed: 0': 'timestamp'})
        elif df_1hour.columns[0] == '' or pd.isna(df_1hour.columns[0]):
            df_1hour = df_1hour.rename(columns={df_1hour.columns[0]: 'timestamp'})

        # Parse timestamps for both dataframes
        if 'timestamp' in df_15min.columns:
            df_15min['timestamp'] = pd.to_datetime(df_15min['timestamp'], utc=True, errors='coerce')
        if 'timestamp' in df_1hour.columns:
            df_1hour['timestamp'] = pd.to_datetime(df_1hour['timestamp'], utc=True, errors='coerce')
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add summary sheet
        ws_summary = wb.create_sheet("Summary", 0)
        
        # Add header
        ws_summary['A1'] = "CEF Butimanu Solar Forecast Report"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A3'] = f"Generated: {datetime.now(pytz.timezone('Europe/Berlin')).strftime('%Y-%m-%d %H:%M:%S CET')}"

        # Add location info
        ws_summary['A5'] = "Location Information"
        ws_summary['A5'].font = Font(bold=True)
        ws_summary['A6'] = "Plant Name:"
        ws_summary['B6'] = "CEF Butimanu"
        ws_summary['A7'] = "DC Capacity:"
        ws_summary['B7'] = "12806 kW (12.806 MW) - LONGi 580W panels"
        ws_summary['A8'] = "AC Capacity:"
        ws_summary['B8'] = "10800 kW (10.8 MW)"
        ws_summary['A9'] = "Location:"
        ws_summary['B9'] = "44.6832°N, 25.9071°E (Butimanu, Dâmbovița)"
        ws_summary['A10'] = "Timezone:"
        ws_summary['B10'] = "CET (Central European Time)"
        
        # Add forecast summary
        ws_summary['A11'] = "Forecast Summary"
        ws_summary['A11'].font = Font(bold=True)
        ws_summary['A12'] = "Forecast Period:"
        ws_summary['B12'] = f"{df_1hour['timestamp'].iloc[0]} to {df_1hour['timestamp'].iloc[-1]}"
        ws_summary['A13'] = "Duration:"
        ws_summary['B13'] = "168 hours (7 days)"
        ws_summary['A14'] = "Peak Power:"
        ws_summary['B14'] = f"{df_1hour['power_kw'].max():.1f} kW"
        ws_summary['A15'] = "Total Energy:"
        ws_summary['B15'] = f"{df_1hour['energy_kwh'].sum():.1f} kWh"
        ws_summary['A16'] = "Average Power:"
        ws_summary['B16'] = f"{df_1hour['power_kw'].mean():.1f} kW"
        
        # Format summary cells
        for row in ws_summary.iter_rows(min_row=1, max_row=16, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(vertical='center')
        
        # Add hourly sheet
        ws_hourly = wb.create_sheet("Hourly Forecast", 1)
        
        # Remove timezone information from datetime columns for Excel compatibility
        for col in df_1hour.columns:
            if pd.api.types.is_datetime64tz_dtype(df_1hour[col]):
                df_1hour[col] = df_1hour[col].dt.tz_localize(None)
        if pd.api.types.is_datetime64tz_dtype(df_1hour.index):
            df_1hour.index = df_1hour.index.tz_localize(None)
        
        # Add header row with formatting
        headers = list(df_1hour.columns)
        ws_hourly.append(headers)
        
        # Format header
        for cell in ws_hourly[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for index, row in df_1hour.iterrows():
            ws_hourly.append(row.tolist())
        
        # Format columns
        for column in ws_hourly.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws_hourly.column_dimensions[column_letter].width = adjusted_width
        
        # Add 15-minute sheet
        ws_15min = wb.create_sheet("15 Minutes", 2)

        # Prepare 15-minute data with time breakdown columns (like hourly)
        df_15min_ordered = df_15min.copy()

        if df_15min_ordered['timestamp'].isna().any():
            raise ValueError("Invalid timestamps detected after parsing")

        # Add time breakdown columns for 15-minute intervals
        df_15min_ordered['YEAR'] = df_15min_ordered['timestamp'].dt.year
        df_15min_ordered['MONTH'] = df_15min_ordered['timestamp'].dt.month
        df_15min_ordered['DAY'] = df_15min_ordered['timestamp'].dt.day
        df_15min_ordered['HOUR_START'] = df_15min_ordered['timestamp'].dt.hour
        df_15min_ordered['MINUTE_START'] = df_15min_ordered['timestamp'].dt.minute

        # Calculate end time (15 minutes later)
        df_15min_ordered['timestamp_end'] = df_15min_ordered['timestamp'] + pd.Timedelta(minutes=15)
        df_15min_ordered['HOUR_END'] = df_15min_ordered['timestamp_end'].dt.hour
        df_15min_ordered['MINUTE_END'] = df_15min_ordered['timestamp_end'].dt.minute
        df_15min_ordered['DAY_END'] = df_15min_ordered['timestamp_end'].dt.day

        # Create interval identifier (e.g., 1-96 for each day)
        df_15min_ordered['interval'] = (df_15min_ordered['HOUR_START'] * 4 +
                                        df_15min_ordered['MINUTE_START'] // 15 + 1)

        # Select and order columns for 15-minute format
        # Show both kW (power) and kWh (energy) for each quartile
        columns_15min = [
            'timestamp', 'YEAR', 'MONTH', 'DAY', 'HOUR_START', 'MINUTE_START',
            'HOUR_END', 'MINUTE_END', 'interval',
            'production_kw', 'energy_kwh',
            'q10', 'energy_q10_kwh',
            'q25', 'energy_q25_kwh',
            'q50', 'energy_q50_kwh',
            'q75', 'energy_q75_kwh',
            'q90', 'energy_q90_kwh'
        ]

        # Add DAY_END only if needed (when interval crosses midnight)
        if (df_15min_ordered['DAY_END'] != df_15min_ordered['DAY']).any():
            columns_15min.insert(8, 'DAY_END')

        # Filter to only existing columns and exclude unwanted ones
        excluded_columns = ['forecast_timestamp', 'resolution_minutes', 'power_kw', 'timestamp_end', 'location']
        columns_15min = [col for col in columns_15min if col in df_15min_ordered.columns and col not in excluded_columns]
        df_15min_ordered = df_15min_ordered[columns_15min]
        
        # Remove timezone information from datetime columns for Excel compatibility
        for col in df_15min_ordered.columns:
            if pd.api.types.is_datetime64tz_dtype(df_15min_ordered[col]):
                df_15min_ordered[col] = df_15min_ordered[col].dt.tz_localize(None)
        
        # Add headers
        headers_15min = list(df_15min_ordered.columns)
        ws_15min.append(headers_15min)
        
        # Format header
        for cell in ws_15min[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for index, row in df_15min_ordered.iterrows():
            ws_15min.append(row.tolist())
        
        # Format columns
        for column in ws_15min.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws_15min.column_dimensions[column_letter].width = adjusted_width
        
        # Generate output filename
        if output_path is None:
            timestamp = datetime.now(pytz.timezone('Europe/Berlin')).strftime('%Y%m%d_%H%M%S')
            output_path = os.path.join(data_dir, f'cef_butimanu_forecast_{timestamp}.xlsx')

        # Save workbook
        wb.save(output_path)
        print(f"\n✅ Excel file created: {output_path}")

        # Display summary
        print("\n📊 Forecast Summary:")
        print(f"   Location: CEF Butimanu (12.806 MW DC / 10.8 MW AC)")
        print(f"   Period: 7 days ({len(df_1hour)} hours)")
        print(f"   Peak Power: {df_1hour['power_kw'].max():.1f} kW")
        print(f"   Total Energy: {df_1hour['energy_kwh'].sum():.1f} kWh")
        print(f"   Data Resolution: 15-minute and hourly")
        
        return output_path
    
    except Exception as e:
        print(f"Error in export_forecast_to_excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    result = export_forecast_to_excel()
    if result is None:
        sys.exit(1)
    sys.exit(0)